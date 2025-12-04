import requests
from openpyxl import load_workbook
import os
from dotenv import load_dotenv
import pandas as pd
from requests.exceptions import RequestException, Timeout, ConnectionError
from openpyxl.styles import PatternFill

load_dotenv()

# Récupération de la clé d'API depuis les variables d'environnement.
# Cela évite de mettre la clé en dur dans le code.
API_KEY = os.getenv('OPENWEATHER_API_KEY')
BASE_URL = os.getenv('OPENWEATHER_API_URL')

# 1. Définir une liste de 10 villes françaises

french_cities_list = [
    "Paris", 
    "Amiens",
    "Lille", 
    "Nice", 
    "Toulouse", 
    "Bordeaux", 
    "Caen",
    "Brest", 
    "Grenoble", 
    "Dijon", 
    "ghdg"
]

# Bonus
def connection(params):
    try:
        response = requests.get(url_meteo, params=params)

        response.raise_for_status()

        return response.json()

    # Timeout : le serveur a mis trop de temps à répondre.
    except Timeout:
        print("Timeout : L'API met trop de temps à répondre")

    # ConnectionError : problème de réseau (pas d'accès internet, DNS, etc.).
    except ConnectionError:
        print("Erreur de connexion : Vérifiez votre réseau ou le serveur")

    # HTTPError : erreur HTTP après raise_for_status() (404, 500, etc.).
    except requests.exceptions.HTTPError as e:
        print(f"Erreur HTTP : {e}")
        # response peut contenir plus d'infos (code + corps de la réponse)
        print(f"Code : {response.status_code}")
        print(f"Message : {response.text}")

    # RequestException : exception générique de requests (base pour toutes les autres).
    except RequestException as e:
        print(f"Erreur générale requests : {e}")

    # ValueError : problème lors du parsing JSON (réponse pas au format JSON).
    except ValueError:
        print("La réponse n'est pas du JSON valide")


# 2. Pour chaque ville, récupérer :
url_meteo = f'{BASE_URL}/weather'

data = []

for city in french_cities_list:
    params = {
        'q': city,
        'appid': API_KEY,
        'units': 'metric',
        'lang': 'fr'
    }
    request = connection(params)
    if request:
        temp_actuelle = request['main']['temp']
        temp_ressentie = request['main']['feels_like']
        humidite = request['main']['humidity']
        description = request['weather'][0]['description']
        data.append({
            'Nom' : city,
            'Température actuelle' : temp_actuelle,
            'Température ressentie' : temp_ressentie,
            'Humidité' : humidite,
            'Description' : description
        })

print(data)

# 3. Créer un DataFrame avec ces informations

df = pd.DataFrame(data)
print(df)

# 4. Identifier la ville la plus chaude et la plus froide

print("# Ville la plus chaude")
hottest_city = df.nlargest(1, "Température actuelle")
print(hottest_city)

print("# Ville la plus froide")
coldest_city = df.nsmallest(1, "Température actuelle")
print(coldest_city)

# 5. Calculer la température moyenne

print("# Température moyenne")
average_temperature = df["Température actuelle"].mean()
print(average_temperature)


# 6. Sauvegarder dans `meteo_villes.csv`

df.to_csv('meteo_villes.csv',index=False)


# bonus export en excel avec 
# couleur des cellules celon la temperature

df.to_excel('meteo_villes.xlsx',index=False)

wb = load_workbook('meteo_villes.xlsx')
ws = wb.active
for row in ws.rows:
    raw_value = row[1].value

    try:
        value = float(raw_value)   # conversion en nombre
    except (TypeError, ValueError):
        continue   # ignore la ligne si ce n'est pas un nombre

    if value < 3:
        fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # light blue
    elif value < 5:
        fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # yellow
    else:
        fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # red

    # appliquer à toute la ligne
    for cell in row:
        cell.fill = fill

wb.save('meteo_villes.xlsx')

